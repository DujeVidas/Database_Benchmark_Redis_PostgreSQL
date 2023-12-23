import openpyxl
import matplotlib.pyplot as plt
import os
import re

sheet_name = "Medians"
current_directory = os.getcwd()
plots_directory = os.path.join(current_directory, "plots")
type_directory = os.path.join(plots_directory, sheet_name)
if not os.path.exists(type_directory):
    os.makedirs(type_directory)


def extract_units(data):
    units = data.split("(")[-1].rstrip(")")
    return units if units else data


def get_data_from_excel(file_path, dataRow, dataColumn, sheet_name):
    
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    cell = sheet.cell(row=dataRow, column=dataColumn)
    data = cell.value
    headCell = sheet.cell(row=1, column=dataColumn)
    headData = headCell.value

    return data, headData


def plot_bar_graph(
    dataRedis,
    dataPostgreSQL,
    file_numbers,
    yLabelRedis,
    yLabelPostgreSQL,
    plotNameRedis,
    plotNamePostgreSQL,
):
    fig, axes = plt.subplots(2, 1, figsize=(8, 10))

    axes[0].bar(range(len(dataRedis)), dataRedis)
    axes[0].set_xlabel("Number Of Iterations")
    axes[0].set_ylabel(yLabelRedis)
    axes[0].set_title(f"Bar Graph of {plotNameRedis}")
    axes[0].set_xticks(range(len(dataRedis)))
    axes[0].set_xticklabels(file_numbers)

    axes[1].bar(range(len(dataPostgreSQL)), dataPostgreSQL)
    axes[1].set_xlabel("Number Of Iterations")
    axes[1].set_ylabel(yLabelPostgreSQL)
    axes[1].set_title(f"Bar Graph of {plotNamePostgreSQL}")
    axes[1].set_xticks(range(len(dataPostgreSQL)))
    axes[1].set_xticklabels(file_numbers)

    fig.tight_layout()
    combined_plot_name = "Redis_vs_" + plotNamePostgreSQL.replace(" ", "_")
    combined_plot_name = combined_plot_name.replace("/", "_")
    save_path = os.path.join(type_directory, f"{combined_plot_name}.png")
    plt.savefig(save_path)
    plt.show()


def numerical_sort(value):
    parts = re.split(r"(\d+)", value)
    parts[1::2] = map(int, parts[1::2])
    return parts


current_directory = os.getcwd()
excel_directory = os.path.join(current_directory, "sheets")
files = [
    f
    for f in os.listdir(excel_directory)
    if os.path.isfile(os.path.join(excel_directory, f))
]
excel_files = [f for f in files if f.endswith(".xlsx")]
excel_files = sorted(excel_files, key=numerical_sort)

all_data_Redis = []
all_data_PostgreSQL = []
file_numbers = []
yLabelRedis = ""
yLabelPostgreSQL = ""
plotNameRedis = ""
plotNamePostgreSQL = ""

for file in excel_files:
    match = re.search(r"\d+", file)
    file_number = match.group() if match else None
    file_numbers.append(file_number)
    file= os.path.join(excel_directory, file)
    excel_data, yLabelData = get_data_from_excel(file, 2, 8, sheet_name)
    yLabelRedis = extract_units(yLabelData)
    all_data_Redis.append(excel_data)
    plotNameRedis = yLabelData
    plotNameRedis = "Redis " + plotNameRedis

    excel_data, yLabelData = get_data_from_excel(file, 3, 8, sheet_name)
    yLabelPostgreSQL = extract_units(yLabelData)
    all_data_PostgreSQL.append(excel_data)
    plotNamePostgreSQL = yLabelData
    plotNamePostgreSQL = "PostgreSQL " + plotNamePostgreSQL

plot_bar_graph(
    all_data_Redis,
    all_data_PostgreSQL,
    file_numbers,
    yLabelRedis,
    yLabelPostgreSQL,
    plotNameRedis,
    plotNamePostgreSQL,
)
