import redis
import psycopg2
from psycopg2 import pool
import pandas as pd
from faker import Faker
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Side, Border
import os
from concurrent.futures import ThreadPoolExecutor
import json
import psutil

NUM_ITERATIONS = 100


current_directory = os.getcwd()
excel_directory = os.path.join(current_directory, "sheets")
if not os.path.exists(excel_directory):
    os.makedirs(excel_directory)

files = [
    f
    for f in os.listdir(excel_directory)
    if os.path.isfile(os.path.join(excel_directory, f))
]
excel_files = [f for f in files if f.endswith(".xlsx")]
fake = Faker()

r = redis.StrictRedis(host="localhost", port=6379, db=0)


connection_pool = pool.SimpleConnectionPool(
    minconn=1,
    maxconn=NUM_ITERATIONS * 10,
    dbname="mydatabase",
    user="dujevidas",
    password="DVidas123",
    host="localhost",
    port="5432",
)


def close_excel_file(file_name):
    """Function to close an Excel file if it's open.

    Args:
        file_name (str): The name of the Excel file to be closed.

    Returns:
        None
    """
    for proc in psutil.process_iter(["pid", "name"]):
        if "EXCEL.EXE" in proc.info["name"]:
            for item in proc.cmdline():
                if file_name in item:
                    print(f"Excel file {file_name} is open. Closing...")
                    os.system(f"TASKKILL /F /PID {proc.pid}")
                    print("Excel file closed.")
                    return
        else:
            print("Excel file is not open.")
            return


if excel_files == []:
    print("No excel files found.")
else:
    for file in excel_files:
        close_excel_file(file)


def redis_read_operation():
    """Function to perform a read operation on a Redis database.

    Args:
        None

    Returns:
        bytes: The value corresponding to the key 'key' in Redis.
    """
    return r.get("key")


def redis_write_operation():
    """Function to perform a write operation on a Redis database.

    Args:
        None

    Returns:
        None
    """
    key = fake.uuid4()
    value = {
        "name": fake.name(),
        "address": fake.address(),
        "phone_number": fake.phone_number(),
        "date_of_birth": fake.date_of_birth().strftime("%Y-%m-%d"),
        "email": fake.email(),
        "credit_card": fake.credit_card_full(),
    }
    r.set(key, json.dumps(value))


def postgres_read_operation():
    """Function to perform a read operation on a PostgreSQL database.

    Args:
        None

    Returns:
        list: A list of tuples containing the data from the 'users' table.
    """
    connection = connection_pool.getconn()
    cursor = connection.cursor()

    cursor.execute("SELECT * FROM users")
    data = cursor.fetchall()

    cursor.close()
    connection_pool.putconn(connection)

    return data


def postgres_write_operation():
    """Function to perform a write operation on a PostgreSQL database.

    Args:
        None

    Returns:
        str: A string indicating the success of the write operation.
    """
    name = fake.name()[:100]
    address = fake.address()
    phone_number = fake.phone_number()[:20]
    date_of_birth = fake.date_of_birth()
    email = fake.email()[:100]
    credit_card = fake.credit_card_full()[:100]
    connection = connection_pool.getconn()
    cursor = connection.cursor()

    cursor.execute(
        "INSERT INTO users (name, address, phone_number, date_of_birth, email, credit_card) VALUES (%s, %s, %s, %s, %s, %s)",
        (name, address, phone_number, date_of_birth, email, credit_card),
    )

    connection.commit()
    cursor.close()
    connection_pool.putconn(connection)

    return "Write successful"


def execute_concurrent_operations_redis():
    """Function to execute concurrent read and write operations on a Redis database.

    Args:
        None

    Returns:
        tuple: A tuple containing the results of the read and write operations.
    """
    with ThreadPoolExecutor(max_workers=10) as executor:
        read_task_redis = executor.submit(redis_read_operation)
        write_task_redis = executor.submit(redis_write_operation)

        read_result_redis = read_task_redis.result()
        write_result_redis = write_task_redis.result()
    return read_result_redis, write_result_redis


def execute_concurrent_operations_postgres():
    """Function to execute concurrent read and write operations on a PostgreSQL database.

    Args:
        None

    Returns:
        tuple: A tuple containing the results of the read and write operations.
    """
    with ThreadPoolExecutor(max_workers=10) as executor:
        read_task_pg = executor.submit(postgres_read_operation)
        write_task_pg = executor.submit(postgres_write_operation)

        read_result_pg = read_task_pg.result()
        write_result_pg = write_task_pg.result()
    return read_result_pg, write_result_pg


concurrent_time_redis = []
for _ in range(NUM_ITERATIONS):
    start_time = time.time()
    execute_concurrent_operations_redis()
    end_time = time.time()
    concurrent_time_redis.append(end_time - start_time)

concurrent_time_postgres = []
for _ in range(NUM_ITERATIONS):
    start_time = time.time()
    execute_concurrent_operations_postgres()
    end_time = time.time()
    concurrent_time_postgres.append(end_time - start_time)


def simulate_read_write_ratio(ratio):
    """Function to simulate a read-write ratio.

    Args:
        ratio (float): The ratio of reads to writes.

    Returns:
        dict: A dictionary containing the results of the simulation.
    """
    num_iterations = 100
    read_count = int(num_iterations * ratio)
    write_count = num_iterations - read_count

    redis_read_times = []
    for _ in range(read_count):
        start_time = time.time()
        redis_read_operation()
        elapsed_time = time.time() - start_time
        redis_read_times.append(elapsed_time)

    redis_write_times = []
    for _ in range(write_count):
        start_time = time.time()
        redis_write_operation()
        elapsed_time = time.time() - start_time
        redis_write_times.append(elapsed_time)

    pg_read_times = []
    for _ in range(read_count):
        start_time = time.time()
        postgres_read_operation()
        elapsed_time = time.time() - start_time
        pg_read_times.append(elapsed_time)

    pg_write_times = []
    for _ in range(write_count):
        start_time = time.time()
        postgres_write_operation()
        elapsed_time = time.time() - start_time
        pg_write_times.append(elapsed_time)

    avg_redis_read_time = sum(redis_read_times) / len(redis_read_times)
    avg_redis_write_time = sum(redis_write_times) / len(redis_write_times)
    avg_pg_read_time = sum(pg_read_times) / len(pg_read_times)
    avg_pg_write_time = sum(pg_write_times) / len(pg_write_times)

    return {
        "Scenario": "Read-Write Ratio",
        "Read Throughput Redis (ops/sec)": read_count / avg_redis_read_time,
        "Write Throughput Redis (ops/sec)": write_count / avg_redis_write_time,
        "Read Throughput PostgreSQL (ops/sec)": read_count / avg_pg_read_time,
        "Write Throughput PostgreSQL (ops/sec)": write_count / avg_pg_write_time,
    }


def simulate_key_access_frequencies():
    """Function to simulate key access frequencies.

    Args:
        None

    Returns:
        dict: A dictionary containing the results of the simulation.
    """
    key_access_scenario_1 = [1, 1, 1, 2, 2, 2, 2, 3, 3, 4]

    redis_access_times = []
    for key_frequency in key_access_scenario_1:
        start_time = time.time()
        for _ in range(key_frequency):
            r.get("key")
        elapsed_time = time.time() - start_time
        redis_access_times.append(elapsed_time)

    pg_access_times = []
    connection = connection_pool.getconn()
    cursor = connection.cursor()
    for key_frequency in key_access_scenario_1:
        start_time = time.time()
        cursor.execute(
            "SELECT * FROM users WHERE key_access_frequency = %s", (key_frequency,)
        )
        rows = cursor.fetchall()
        elapsed_time = time.time() - start_time
        pg_access_times.append(elapsed_time)
    cursor.close()
    connection_pool.putconn(connection)
    avg_redis_access_time = sum(redis_access_times) / len(redis_access_times)
    avg_pg_access_time = sum(pg_access_times) / len(pg_access_times)

    return {
        "Redis Avg Response Time (sec)": avg_redis_access_time,
        "PostgreSQL Avg Response Time (sec)": avg_pg_access_time,
    }


def simulate_redis_row_load():
    """Function to simulate row load on Redis.

    Args:
        None

    Returns:
        dict: A dictionary containing the results of the simulation.
    """
    rows_queried = [10, 20, 30, 40, 50]

    redis_access_times = []
    for rows in rows_queried:
        start_time = time.time()
        for _ in range(rows):
            r.get("key")
        elapsed_time = time.time() - start_time
        redis_access_times.append(elapsed_time)

    avg_redis_access_time = sum(redis_access_times) / len(redis_access_times)

    return {
        "Scenario": "Redis Row Load",
        "Redis Avg Row Load Response Time (sec)": avg_redis_access_time,
    }


def simulate_postgres_row_load():
    """Function to simulate row load on PostgreSQL.

    Args:
        None

    Returns:
        dict: A dictionary containing the results of the simulation.
    """
    rows_queried = [10, 20, 30, 40, 50]

    pg_access_times = []
    connection = connection_pool.getconn()
    cursor = connection.cursor()
    for rows in rows_queried:
        start_time = time.time()
        cursor.execute("SELECT * FROM users LIMIT %s", (rows,))
        rows = cursor.fetchall()
        elapsed_time = time.time() - start_time
        pg_access_times.append(elapsed_time)
    cursor.close()
    connection_pool.putconn(connection)
    avg_pg_access_time = sum(pg_access_times) / len(pg_access_times)

    return {
        "Scenario": "PostgreSQL Row Load",
        "PostgreSQL Avg Row Load Response Time (sec)": avg_pg_access_time,
    }


def read_heavy_transaction_redis(num_reads=10):
    """
    This function simulates a read-heavy transaction on Redis.

    Args:
        num_reads (int): The number of reads to simulate.

    Returns:
        float: The time it took to complete the transaction.
    """
    start_time = time.time()

    cursor = 0
    for _ in range(num_reads):
        keys = []
        while True:
            cursor, partial_keys = r.scan(cursor, match="A*", count=1000)
            keys.extend(partial_keys)
            if cursor == 0:
                break

        for key in keys:
            r.get(key)

    end_time = time.time()
    return end_time - start_time


def read_heavy_transaction_postgres(num_reads=10):
    """
    This function simulates a read-heavy transaction on PostgreSQL.

    Args:
        num_reads (int): The number of reads to simulate.

    Returns:
        float: The time it took to complete the transaction.
    """
    start_time = time.time()
    connection = connection_pool.getconn()
    cursor = connection.cursor()

    for _ in range(num_reads):
        cursor.execute("SELECT * FROM users WHERE name LIKE 'A%'")
        cursor.fetchall()

    cursor.close()
    connection_pool.putconn(connection)
    end_time = time.time()
    return end_time - start_time


def write_heavy_transaction_redis(num_writes=10):
    """
    This function simulates a write-heavy transaction on Redis.

    Args:
        num_writes (int): The number of writes to simulate.

    Returns:
        float: The time it took to complete the transaction.
    """

    start_time = time.time()

    for _ in range(num_writes):
        key = fake.uuid4()
        value = {
            "name": fake.first_name(),
            "address": fake.address(),
            "phone_number": fake.phone_number(),
            "date_of_birth": fake.date_of_birth().strftime("%Y-%m-%d"),
            "email": fake.email(),
            "credit_card": fake.credit_card_full(),
        }
        r.set(key, json.dumps(value))

    end_time = time.time()
    return end_time - start_time


def write_heavy_transaction_postgres(num_writes=10):
    """
    This function simulates a write-heavy transaction on PostgreSQL.

    Args:
        num_writes (int): The number of writes to simulate.

    Returns:
        float: The time it took to complete the transaction.
    """
    start_time = time.time()

    connection = connection_pool.getconn()
    cursor = connection.cursor()

    for _ in range(num_writes):
        cursor.execute(
            "INSERT INTO users (name, address, phone_number, date_of_birth, email, credit_card) VALUES (%s, %s, %s, %s, %s, %s)",
            (
                fake.first_name()[:100],
                fake.address(),
                fake.phone_number()[:20],
                fake.date_of_birth(),
                fake.email()[:100],
                fake.credit_card_full()[:100],
            ),
        )

    connection.commit()
    cursor.close()
    connection_pool.putconn(connection)

    end_time = time.time()
    return end_time - start_time


def pg_perform_transactional_operations():
    """
    This function performs transactional operations on PostgreSQL.

    Args:
        None

    Returns:
        str: The result of the transaction.
    """
    transaction_results = ""
    connection = connection_pool.getconn()
    try:
        with connection:
            with connection.cursor() as cursor:
                cursor.execute("SELECT * FROM users WHERE name LIKE 'A%'")
                rows = cursor.fetchall()

                cursor.execute(
                    "INSERT INTO users (name, address, phone_number, date_of_birth, email, credit_card) VALUES (%s, %s, %s, %s, %s, %s)",
                    (
                        fake.first_name()[:100],
                        fake.address(),
                        fake.phone_number()[:20],
                        fake.date_of_birth(),
                        fake.email()[:100],
                        fake.credit_card_full()[:100],
                    ),
                )

                cursor.execute(
                    "UPDATE users SET email = %s WHERE name LIKE 'A%%'",
                    (fake.email()[:100],),
                )

            connection.commit()
            transaction_results = "Commited"

    except psycopg2.DatabaseError as e:
        connection.rollback()
        transaction_results = "Rolled back"

    finally:
        connection.close()

    return transaction_results


def redis_perform_transactional_operations():
    """
    This function performs transactional operations on Redis.

    Args:
        None

    Returns:
        str: The result of the transaction.
    """
    transaction_results = ""
    try:
        keys_starting_with_a = [key for key in r.keys() if key.startswith(b"A")]

        for key in keys_starting_with_a:
            value = r.get(key)

        redis_write_operation()

        for key in keys_starting_with_a:
            new_value = {
                "name": fake.name(),
                "address": fake.address(),
                "phone_number": fake.phone_number(),
                "date_of_birth": fake.date_of_birth().strftime("%Y-%m-%d"),
                "email": fake.email(),
                "credit_card": fake.credit_card_full(),
            }
            r.set(key, new_value)

        transaction_results = "Commited"

    except Exception as e:
        transaction_results = "Rolled Back"

    return transaction_results


def colourSheet(sheet):
    """
    This function colourizes the given sheet based on the values in the specified columns.

    Args:
        sheet (Worksheet): The sheet to colourize.

    Returns:
        None
    """
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    for col in range(2, 4):
        column_values = [
            sheet.cell(row=row, column=col).value for row in range(2, num_rows)
        ]
        min_value = min(column_values)
        max_value = max(column_values)

        for row in range(2, num_rows):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value

            if cell_value == max_value:
                cell.fill = PatternFill(
                    start_color="00FF00", end_color="00FF00", fill_type="solid"
                )
            elif cell_value == min_value:
                cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )
            bStyle = "thin"
            cell.border = Border(
                top=Side(border_style=bStyle, color="000000"),
                left=Side(border_style=bStyle, color="000000"),
                right=Side(border_style=bStyle, color="000000"),
                bottom=Side(border_style=bStyle, color="000000"),
            )

    for col in range(4, num_cols + 1):
        column_values = [
            sheet.cell(row=row, column=col).value for row in range(2, num_rows)
        ]
        min_value = min(column_values)
        max_value = max(column_values)

        for row in range(2, num_rows):
            cell = sheet.cell(row=row, column=col)
            cell_value = cell.value

            if cell_value == min_value:
                cell.fill = PatternFill(
                    start_color="00FF00", end_color="00FF00", fill_type="solid"
                )
            elif cell_value == max_value:
                cell.fill = PatternFill(
                    start_color="FF0000", end_color="FF0000", fill_type="solid"
                )

            bStyle = "thin"
            cell.border = Border(
                top=Side(border_style=bStyle, color="000000"),
                left=Side(border_style=bStyle, color="000000"),
                right=Side(border_style=bStyle, color="000000"),
                bottom=Side(border_style=bStyle, color="000000"),
            )
    for col in range(2, num_cols + 1):
        cell = sheet.cell(row=num_rows, column=col)
        cell.fill = PatternFill(
            start_color="0000FF", end_color="0000FF", fill_type="solid"
        )
        bStyle = "thick"
        cell.border = Border(
            top=Side(border_style=bStyle, color="000000"),
            left=Side(border_style=bStyle, color="000000"),
            right=Side(border_style=bStyle, color="000000"),
            bottom=Side(border_style=bStyle, color="000000"),
        )


redis_read_results = []
redis_write_results = []
pg_read_results = []
pg_write_results = []

key_access_results = []
redis_response_times = []
pg_response_times = []

redis_row_load_results = []
pg_row_load_results = []

redis_read_heavy_transaction_results = []
pg_read_heavy_transaction_results = []

redis_write_heavy_transaction_results = []
pg_write_heavy_transaction_results = []

redis_transactional_operations_results = []
pg_transactional_operations_results = []

"""
    This is the main loop that runs the benchmark.
    It iterates X times, simulating the following:
        - Simulating read and write ratios
        - Simulating key access frequencies
        - Simulating row load times
        - Simulating read heavy transactions
        - Simulating write heavy transactions
        - Simulating transactional operations
"""
for _ in range(NUM_ITERATIONS):
    redis_read_write_metrics = simulate_read_write_ratio(0.7)
    redis_read_results.append(
        redis_read_write_metrics["Read Throughput Redis (ops/sec)"]
    )
    redis_write_results.append(
        redis_read_write_metrics["Write Throughput Redis (ops/sec)"]
    )

    pg_read_write_metrics = simulate_read_write_ratio(0.7)
    pg_read_results.append(
        pg_read_write_metrics["Read Throughput PostgreSQL (ops/sec)"]
    )
    pg_write_results.append(
        pg_read_write_metrics["Write Throughput PostgreSQL (ops/sec)"]
    )

    key_access_metrics = simulate_key_access_frequencies()
    key_access_results.append(key_access_metrics)
    redis_response_times.append(key_access_metrics["Redis Avg Response Time (sec)"])
    pg_response_times.append(key_access_metrics["PostgreSQL Avg Response Time (sec)"])

    redis_row_load_metrics = simulate_redis_row_load()
    redis_row_load_results.append(
        redis_row_load_metrics["Redis Avg Row Load Response Time (sec)"]
    )

    pg_row_load_metrics = simulate_postgres_row_load()
    pg_row_load_results.append(
        pg_row_load_metrics["PostgreSQL Avg Row Load Response Time (sec)"]
    )

    redis_read_heavy_transaction_results.append(read_heavy_transaction_redis())
    pg_read_heavy_transaction_results.append(read_heavy_transaction_postgres())

    redis_write_heavy_transaction_results.append(write_heavy_transaction_redis())
    pg_write_heavy_transaction_results.append(write_heavy_transaction_postgres())

    redis_transactional_operations_results.append(
        redis_perform_transactional_operations()
    )
    pg_transactional_operations_results.append(pg_perform_transactional_operations())

"""
    This is the code that generates the Excel file.
    It creates a new Excel file, populates it with the data from the benchmark,
    and colourizes the cells based on the values in the specified columns.
    Finally, it saves the Excel file to disk.
"""

df_redis_times_concurrent = pd.DataFrame(
    {"Concurrent Time Redis (seconds)": concurrent_time_redis}
)
df_postgres_times_concurrent = pd.DataFrame(
    {"Concurrent Time PostgreSQL (seconds)": concurrent_time_postgres}
)


df_key_access = pd.DataFrame(key_access_results)
avg_redis_response_time = sum(redis_response_times) / len(redis_response_times)
avg_pg_response_time = sum(pg_response_times) / len(pg_response_times)

df_redis = pd.DataFrame(
    {
        "Read Throughput Redis (ops/sec)": redis_read_results,
        "Write Throughput Redis (ops/sec)": redis_write_results,
    }
)

df_pg = pd.DataFrame(
    {
        "Read Throughput PostgreSQL (ops/sec)": pg_read_results,
        "Write Throughput PostgreSQL (ops/sec)": pg_write_results,
    }
)

df_redis_row_load = pd.DataFrame(
    {"Redis Avg Row Load Response Time (sec)": redis_row_load_results}
)

df_pg_row_load = pd.DataFrame(
    {"PostgreSQL Avg Row Load Response Time (sec)": pg_row_load_results}
)

df_redis_read_heavy_transaction = pd.DataFrame(
    {
        "Redis Read Heavy Transaction Time (seconds)": redis_read_heavy_transaction_results
    }
)

df_pg_read_heavy_transaction = pd.DataFrame(
    {
        "PostgreSQL Read Heavy Transaction Time (seconds)": pg_read_heavy_transaction_results
    }
)

df_redis_write_heavy_transaction = pd.DataFrame(
    {
        "Redis Write Heavy Transaction Time (seconds)": redis_write_heavy_transaction_results
    }
)

df_pg_write_heavy_transaction = pd.DataFrame(
    {
        "PostgreSQL Write Heavy Transaction Time (seconds)": pg_write_heavy_transaction_results
    }
)

df_redis_transactional_operations = pd.DataFrame(
    {"Redis Transactional Operations": redis_transactional_operations_results}
)

df_pg_transactional_operations = pd.DataFrame(
    {"PostgreSQL Transactional Operations": pg_transactional_operations_results}
)

avg_redis_read_throughput = df_redis["Read Throughput Redis (ops/sec)"].mean()
avg_redis_write_throughput = df_redis["Write Throughput Redis (ops/sec)"].mean()
avg_pg_read_throughput = df_pg["Read Throughput PostgreSQL (ops/sec)"].mean()
avg_pg_write_throughput = df_pg["Write Throughput PostgreSQL (ops/sec)"].mean()
avg_redis_row_query_time = df_redis_row_load[
    "Redis Avg Row Load Response Time (sec)"
].mean()
avg_pg_row_query_time = df_pg_row_load[
    "PostgreSQL Avg Row Load Response Time (sec)"
].mean()
avg_redis_concurent_time = df_redis_times_concurrent[
    "Concurrent Time Redis (seconds)"
].mean()
avg_pg_concurent_time = df_postgres_times_concurrent[
    "Concurrent Time PostgreSQL (seconds)"
].mean()
avg_redis_read_heavy_transaction = df_redis_read_heavy_transaction[
    "Redis Read Heavy Transaction Time (seconds)"
].mean()
avg_pg_read_heavy_transaction = df_pg_read_heavy_transaction[
    "PostgreSQL Read Heavy Transaction Time (seconds)"
].mean()
avg_redis_write_heavy_transaction = df_redis_write_heavy_transaction[
    "Redis Write Heavy Transaction Time (seconds)"
].mean()
avg_pg_write_heavy_transaction = df_pg_write_heavy_transaction[
    "PostgreSQL Write Heavy Transaction Time (seconds)"
].mean()

median_redis_read_throughput = df_redis["Read Throughput Redis (ops/sec)"].median()
median_redis_write_throughput = df_redis["Write Throughput Redis (ops/sec)"].median()
median_pg_read_throughput = df_pg["Read Throughput PostgreSQL (ops/sec)"].median()
median_pg_write_throughput = df_pg["Write Throughput PostgreSQL (ops/sec)"].median()
median_redis_response_time = df_key_access["Redis Avg Response Time (sec)"].median()
median_pg_response_time = df_key_access["PostgreSQL Avg Response Time (sec)"].median()
median_redis_row_query_time = df_redis_row_load[
    "Redis Avg Row Load Response Time (sec)"
].median()
median_pg_row_query_time = df_pg_row_load[
    "PostgreSQL Avg Row Load Response Time (sec)"
].median()
median_redis_concurent_time = df_redis_times_concurrent[
    "Concurrent Time Redis (seconds)"
].median()
median_pg_concurent_time = df_postgres_times_concurrent[
    "Concurrent Time PostgreSQL (seconds)"
].median()
median_redis_read_heavy_transaction = df_redis_read_heavy_transaction[
    "Redis Read Heavy Transaction Time (seconds)"
].median()
median_pg_read_heavy_transaction = df_pg_read_heavy_transaction[
    "PostgreSQL Read Heavy Transaction Time (seconds)"
].median()
median_redis_write_heavy_transaction = df_redis_write_heavy_transaction[
    "Redis Write Heavy Transaction Time (seconds)"
].median()
median_pg_write_heavy_transaction = df_pg_write_heavy_transaction[
    "PostgreSQL Write Heavy Transaction Time (seconds)"
].median()

df_averages = pd.DataFrame(
    {
        "Scenario": ["Redis", "PostgreSQL", "Difference"],
        "Average Read Throughput (ops/sec)": [
            avg_redis_read_throughput,
            avg_pg_read_throughput,
            abs(avg_redis_read_throughput - avg_pg_read_throughput),
        ],
        "Average Write Throughput (ops/sec)": [
            avg_redis_write_throughput,
            avg_pg_write_throughput,
            abs(avg_redis_write_throughput - avg_pg_write_throughput),
        ],
        "Average Response Time - KAF (sec)": [
            avg_redis_response_time,
            avg_pg_response_time,
            abs(avg_redis_response_time - avg_pg_response_time),
        ],
        "Average Row Query Time (sec)": [
            avg_redis_row_query_time,
            avg_pg_row_query_time,
            abs(avg_redis_row_query_time - avg_pg_row_query_time),
        ],
        "Average Concurrent Time (seconds)": [
            avg_redis_concurent_time,
            avg_pg_concurent_time,
            abs(avg_redis_concurent_time - avg_pg_concurent_time),
        ],
        "Average Read Heavy Transaction Time (seconds)": [
            avg_redis_read_heavy_transaction,
            avg_pg_read_heavy_transaction,
            abs(avg_redis_read_heavy_transaction - avg_pg_read_heavy_transaction),
        ],
        "Average Write Heavy Transaction Time (seconds)": [
            avg_redis_write_heavy_transaction,
            avg_pg_write_heavy_transaction,
            abs(avg_redis_write_heavy_transaction - avg_pg_write_heavy_transaction),
        ],
    }
)

df_medians = pd.DataFrame(
    {
        "Scenario": ["Redis", "PostgreSQL", "Difference"],
        "Median Read Throughput (ops/sec)": [
            median_redis_read_throughput,
            median_pg_read_throughput,
            abs(median_redis_read_throughput - median_pg_read_throughput),
        ],
        "Median Write Throughput (ops/sec)": [
            median_redis_write_throughput,
            median_pg_write_throughput,
            abs(median_redis_write_throughput - median_pg_write_throughput),
        ],
        "Median Response Time - KAF (sec)": [
            median_redis_response_time,
            median_pg_response_time,
            abs(median_redis_response_time - median_pg_response_time),
        ],
        "Median Row Query Time (sec)": [
            median_redis_row_query_time,
            median_pg_row_query_time,
            abs(median_redis_row_query_time - median_pg_row_query_time),
        ],
        "Median Concurrent Time (seconds)": [
            median_redis_concurent_time,
            median_pg_concurent_time,
            abs(median_redis_concurent_time - median_pg_concurent_time),
        ],
        "Median Read Heavy Transaction Time (seconds)": [
            median_redis_read_heavy_transaction,
            median_pg_read_heavy_transaction,
            abs(median_redis_read_heavy_transaction - median_pg_read_heavy_transaction),
        ],
        "Median Write Heavy Transaction Time (seconds)": [
            median_redis_write_heavy_transaction,
            median_pg_write_heavy_transaction,
            abs(
                median_redis_write_heavy_transaction - median_pg_write_heavy_transaction
            ),
        ],
    }
)

df_redis_row_load = pd.DataFrame(
    {"Redis Avg Row Load Response Time (sec)": redis_row_load_results}
)

df_pg_row_load = pd.DataFrame(
    {"PostgreSQL Avg Row Load Response Time (sec)": pg_row_load_results}
)


# Creating an Excel file
myWorkbook = os.path.join(
    excel_directory, "database_metrics_{}.xlsx".format(NUM_ITERATIONS)
)
pd.set_option("display.max_colwidth", None)

# Writing to Excel
with pd.ExcelWriter(myWorkbook) as writer:
    df_averages.to_excel(writer, sheet_name="Averages", index=False)
    df_medians.to_excel(writer, sheet_name="Medians", index=False)

    df_redis.to_excel(writer, sheet_name="Redis_Read_Write_Ratio", index=False)
    df_pg.to_excel(writer, sheet_name="PostgreSQL_Read_Write_Ratio", index=False)

    df_key_access.to_excel(writer, sheet_name="Key_Access_Frequencies", index=False)

    df_redis_row_load.to_excel(writer, sheet_name="Redis_Row_Load", index=False),
    df_pg_row_load.to_excel(writer, sheet_name="PostgreSQL_Row_Load", index=False),

    df_redis_times_concurrent.to_excel(
        writer, sheet_name="Redis_Concurrent_Times", index=False
    ),
    df_postgres_times_concurrent.to_excel(
        writer, sheet_name="PostgreSQL_Concurrent_Times", index=False
    )

    df_redis_read_heavy_transaction.to_excel(
        writer, sheet_name="Redis Read Heavy Trans", index=False
    )
    df_pg_read_heavy_transaction.to_excel(
        writer, sheet_name="PostgreSQL Read Heavy Trans", index=False
    )

    df_redis_write_heavy_transaction.to_excel(
        writer, sheet_name="Redis Write Heavy Trans", index=False
    )
    df_pg_write_heavy_transaction.to_excel(
        writer, sheet_name="PostgreSQL Write Heavy Trans", index=False
    )

    df_redis_transactional_operations.to_excel(
        writer, sheet_name="Redis Trans Operations", index=False
    )
    df_pg_transactional_operations.to_excel(
        writer, sheet_name="PostgreSQL Trans Operations", index=False
    )


workbook = load_workbook(myWorkbook)
sheet_averages = workbook["Averages"]
sheet_medians = workbook["Medians"]
# Colouring Sheets That Contain Averages and Medians
colourSheet(sheet_averages)
colourSheet(sheet_medians)
# Expanding Cells to Fit Data
for sheet in workbook.sheetnames:
    ws = workbook[sheet]
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

# Save & Open
workbook.save(myWorkbook)
os.system('start excel.exe "{}"'.format(myWorkbook))
